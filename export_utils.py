import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from extensions import db
from models import Store, Employee, Schedule

HEADER_FONT = Font(bold=True, size=12, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='8EA9DB', end_color='8EA9DB', fill_type='solid')
WARN_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
SENIOR_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def _apply_style(cell, font=None, fill=None, align=None, border=True):
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = THIN_BORDER

def _parse(d):
    return datetime.strptime(d, '%Y-%m-%d').date()

def _time_slots(store, schedules):
    open_t = datetime.strptime(store.open_time, '%H:%M')
    close_t = datetime.strptime(store.close_time, '%H:%M')
    slots = []
    cur = open_t
    while cur < close_t:
        nxt = cur + timedelta(hours=1)
        key = cur.strftime('%H:%M')
        staff = []
        for s in schedules:
            ss = datetime.strptime(s.start_time, '%H:%M')
            se = datetime.strptime(s.end_time, '%H:%M')
            if ss <= cur and se > cur:
                staff.append({
                    'schedule_id': s.id,
                    'name': s.employee.name,
                    'skill_level': s.employee.skill_level
                })
        slots.append({
            'time': key,
            'staff': staff,
            'count': len(staff),
            'senior_count': len([s for s in staff if s['skill_level'] == '高级']),
            'meets_minimum': len(staff) >= store.min_staff,
            'shortage': max(0, store.min_staff - len(staff))
        })
        cur = nxt
    return slots

def _write_summary_sheet(wb, start_date, end_date, stores):
    ws = wb.create_sheet('排班汇总', 0)
    ws.merge_cells('A1:I1')
    ws['A1'] = f'员工排班汇总表 ({start_date.isoformat()} 至 {end_date.isoformat()})'
    _apply_style(ws['A1'], font=Font(bold=True, size=16), align=CENTER, border=False)
    ws.row_dimensions[1].height = 30

    headers = ['门店', '日期', '星期', '时段', '在岗人员', '人数', '高级数', '达标', '备注']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        _apply_style(c, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)

    row = 4
    cur_date = start_date
    while cur_date <= end_date:
        day_name = ['周一','周二','周三','周四','周五','周六','周日'][cur_date.weekday()]
        for store in stores:
            scheds = Schedule.query.filter_by(store_id=store.id, date=cur_date).all()
            slots = _time_slots(store, scheds)
            for i, slot in enumerate(slots):
                names = '、'.join([s['name'] for s in slot['staff']]) or '-'
                ws.cell(row=row, column=1, value=store.name)
                ws.cell(row=row, column=2, value=cur_date.isoformat())
                ws.cell(row=row, column=3, value=day_name)
                ws.cell(row=row, column=4, value=slot['time'])
                ws.cell(row=row, column=5, value=names)
                ws.cell(row=row, column=6, value=slot['count'])
                ws.cell(row=row, column=7, value=slot['senior_count'])
                c8 = ws.cell(row=row, column=8, value='是' if slot['meets_minimum'] else '否')
                if not slot['meets_minimum']:
                    _apply_style(c8, fill=WARN_FILL)
                remark = ''
                if not slot['meets_minimum']:
                    remark = f'缺{slot["shortage"]}人'
                if slot['count'] > 0 and slot['senior_count'] == 0:
                    remark = (remark + ' ' if remark else '') + '无高级员工'
                ws.cell(row=row, column=9, value=remark)
                for col in range(1, 10):
                    _apply_style(ws.cell(row=row, column=col), align=CENTER)
                row += 1
        cur_date += timedelta(days=1)

    widths = [10, 13, 7, 8, 38, 7, 8, 7, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _write_store_sheet(wb, store, start_date, end_date, detailed=True):
    ws = wb.create_sheet(f'{store.name}排班')
    ws.merge_cells('A1:J1')
    ws['A1'] = f'{store.name} 排班表 ({start_date.isoformat()} 至 {end_date.isoformat()})'
    _apply_style(ws['A1'], font=Font(bold=True, size=15), align=CENTER, border=False)
    ws.row_dimensions[1].height = 28

    ws['A2'] = f'营业时间 {store.open_time}-{store.close_time}  |  最低在岗 {store.min_staff} 人  |  地址：{store.address}'
    _apply_style(ws['A2'], font=Font(italic=True, size=10), align=CENTER, border=False)
    ws.merge_cells('A2:J2')

    date_headers = ['时段/人员']
    cur = start_date
    while cur <= end_date:
        dn = ['一','二','三','四','五','六','日'][cur.weekday()]
        date_headers.append(f'{cur.month}/{cur.day}\n周{dn}')
        cur += timedelta(days=1)

    for col, h in enumerate(date_headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        _apply_style(c, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        ws.row_dimensions[4].height = 30

    open_t = datetime.strptime(store.open_time, '%H:%M')
    close_t = datetime.strptime(store.close_time, '%H:%M')
    time_rows = []
    t = open_t
    while t < close_t:
        time_rows.append(t.strftime('%H:%M'))
        t += timedelta(hours=1)

    for ri, tm in enumerate(time_rows, 5):
        ws.cell(row=ri, column=1, value=tm)
        _apply_style(ws.cell(row=ri, column=1), align=CENTER)

        cur_date = start_date
        for ci in range(2, len(date_headers) + 1):
            scheds = Schedule.query.filter_by(store_id=store.id, date=cur_date).all()
            cur_time = datetime.strptime(tm, '%H:%M')
            staff = []
            for s in scheds:
                ss = datetime.strptime(s.start_time, '%H:%M')
                se = datetime.strptime(s.end_time, '%H:%M')
                if ss <= cur_time and se > cur_time:
                    staff.append(s.employee)

            if staff:
                text = '\n'.join([f"{e.name}({'高' if e.skill_level == '高级' else '初'})" for e in staff])
            else:
                text = ''
            cell = ws.cell(row=ri, column=ci, value=text)
            _apply_style(cell, align=CENTER)
            if len(staff) < store.min_staff:
                _apply_style(cell, fill=WARN_FILL)
            elif any(e.skill_level == '高级' for e in staff) and len(staff) >= store.min_staff:
                pass
            ws.row_dimensions[ri].height = 35
            cur_date += timedelta(days=1)

    ws.column_dimensions['A'].width = 10
    for ci in range(2, len(date_headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 15

    if detailed:
        emp_row = len(time_rows) + 7
        ws.merge_cells(f'A{emp_row}:J{emp_row}')
        ws.cell(row=emp_row, column=1, value='员工明细（按日排班时间）')
        _apply_style(ws.cell(row=emp_row, column=1), font=Font(bold=True, size=12), align=CENTER, border=False)

        employees = Employee.query.all()
        header_row = emp_row + 2
        hs = ['员工'] + date_headers[1:] + ['总工时']
        for ci, h in enumerate(hs, 1):
            c = ws.cell(row=header_row, column=ci, value=h)
            _apply_style(c, font=HEADER_FONT, fill=SUBHEADER_FILL, align=CENTER)

        for ri, emp in enumerate(employees, header_row + 1):
            cell = ws.cell(row=ri, column=1, value=f'{emp.name}({emp.skill_level})')
            _apply_style(cell, align=CENTER)
            if emp.skill_level == '高级':
                _apply_style(cell, fill=SENIOR_FILL)

            total_h = 0
            cur_date = start_date
            for ci in range(2, len(date_headers) + 1):
                sched = Schedule.query.filter_by(
                    employee_id=emp.id, store_id=store.id, date=cur_date
                ).first()
                if sched:
                    val = f'{sched.start_time}-{sched.end_time}'
                    dur = sched._get_duration()
                    total_h += dur
                    c = ws.cell(row=ri, column=ci, value=val)
                    if emp.skill_level == '高级':
                        _apply_style(c, fill=SENIOR_FILL)
                else:
                    c = ws.cell(row=ri, column=ci, value='')
                _apply_style(c, align=CENTER)
                cur_date += timedelta(days=1)

            c = ws.cell(row=ri, column=len(hs), value=round(total_h, 1))
            _apply_style(c, align=CENTER)

def _write_workhours_sheet(wb, start_date, end_date):
    ws = wb.create_sheet('工时统计')
    ws.merge_cells('A1:H1')
    ws['A1'] = f'员工工时统计 ({start_date.isoformat()} 至 {end_date.isoformat()})'
    _apply_style(ws['A1'], font=Font(bold=True, size=14), align=CENTER, border=False)
    ws.row_dimensions[1].height = 25

    headers = ['员工', '技能', '排班天数', '总工时', '日均工时', '标准(月)', '状态', '超时']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        _apply_style(c, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)

    employees = Employee.query.all()
    day_count = (end_date - start_date).days + 1
    for ri, emp in enumerate(employees, 4):
        scheds = Schedule.query.filter(
            Schedule.employee_id == emp.id,
            Schedule.date >= start_date,
            Schedule.date <= end_date
        ).all()
        total_h = round(sum(s._get_duration() for s in scheds), 1)
        days = len(set(s.date for s in scheds))
        avg = round(total_h / days, 1) if days else 0
        std = 160
        is_over = total_h > std
        over = round(max(0, total_h - std), 1)

        ws.cell(row=ri, column=1, value=emp.name)
        ws.cell(row=ri, column=2, value=emp.skill_level)
        ws.cell(row=ri, column=3, value=days)
        ws.cell(row=ri, column=4, value=total_h)
        ws.cell(row=ri, column=5, value=avg)
        ws.cell(row=ri, column=6, value=std)
        c7 = ws.cell(row=ri, column=7, value='超时' if is_over else '正常')
        if is_over:
            _apply_style(c7, fill=WARN_FILL)
        ws.cell(row=ri, column=8, value=over)
        for col in range(1, 9):
            _apply_style(ws.cell(row=ri, column=col), align=CENTER)

    widths = [15, 8, 10, 10, 10, 10, 8, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def export_to_excel(start_date_str, end_date_str, store_id=None, mode='full'):
    start_date = _parse(start_date_str)
    end_date = _parse(end_date_str)

    if store_id:
        stores = [Store.query.get(store_id)]
        stores = [s for s in stores if s]
    else:
        stores = Store.query.all()

    wb = Workbook()
    wb.remove(wb.active)

    if mode == 'simple':
        for s in stores:
            _write_simple_week_sheet(wb, s, start_date, end_date)
    elif mode == 'week' or mode == 'full':
        _write_summary_sheet(wb, start_date, end_date, stores)
        for s in stores:
            _write_store_sheet(wb, s, start_date, end_date, detailed=True)
        if not store_id:
            _write_workhours_sheet(wb, start_date, end_date)
    else:
        _write_summary_sheet(wb, start_date, end_date, stores)
        for s in stores:
            _write_store_sheet(wb, s, start_date, end_date, detailed=True)
        if not store_id:
            _write_workhours_sheet(wb, start_date, end_date)

    file_path = os.path.join(os.path.dirname(__file__), 'temp_schedule.xlsx')
    wb.save(file_path)
    return file_path

def _write_simple_week_sheet(wb, store, start_date, end_date):
    ws = wb.create_sheet(f'{store.name}_本周排班')
    
    day_count = (end_date - start_date).days + 1
    days = []
    cur = start_date
    while cur <= end_date:
        day_name = ['周一','周二','周三','周四','周五','周六','周日'][cur.weekday()]
        days.append((cur, f'{cur.month}/{cur.day}\n{day_name}'))
        cur += timedelta(days=1)
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + day_count)
    ws.cell(row=1, column=1, value=f'{store.name} 本周排班表 ({start_date.isoformat()} ~ {end_date.isoformat()})')
    _apply_style(ws.cell(row=1, column=1), font=Font(bold=True, size=14), align=CENTER, border=False)
    ws.row_dimensions[1].height = 28
    
    headers = ['员工', '技能'] + [d[1] for d in days] + ['本周工时']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        _apply_style(c, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    ws.row_dimensions[3].height = 32
    
    emp_scheds = {}
    cur = start_date
    while cur <= end_date:
        scheds = Schedule.query.filter_by(store_id=store.id, date=cur).all()
        for s in scheds:
            if s.employee_id not in emp_scheds:
                emp_scheds[s.employee_id] = {'name': s.employee.name, 'skill': s.employee.skill_level, 'days': {}}
            emp_scheds[s.employee_id]['days'][cur.isoformat()] = f'{s.start_time}-{s.end_time}'
        cur += timedelta(days=1)
    
    emp_list = sorted(emp_scheds.values(), key=lambda x: (0 if x['skill'] == '高级' else 1, x['name']))
    
    row = 4
    for emp in emp_list:
        ws.cell(row=row, column=1, value=emp['name'])
        skill_cell = ws.cell(row=row, column=2, value=emp['skill'])
        if emp['skill'] == '高级':
            _apply_style(skill_cell, fill=SENIOR_FILL, align=CENTER)
        else:
            _apply_style(skill_cell, align=CENTER)
        
        total_hours = 0
        for i, (date, _) in enumerate(days):
            date_str = date.isoformat()
            shift = emp['days'].get(date_str, '')
            cell = ws.cell(row=row, column=3 + i, value=shift)
            if shift:
                start_t = datetime.strptime(shift.split('-')[0], '%H:%M')
                end_t = datetime.strptime(shift.split('-')[1], '%H:%M')
                total_hours += (end_t - start_t).total_seconds() / 3600
            _apply_style(cell, align=CENTER)
        
        ws.cell(row=row, column=3 + day_count, value=round(total_hours, 1))
        _apply_style(ws.cell(row=row, column=1), align=CENTER)
        _apply_style(ws.cell(row=row, column=3 + day_count), align=CENTER)
        row += 1
    
    ws.cell(row=row, column=1, value='在岗人数')
    ws.cell(row=row, column=2, value='')
    _apply_style(ws.cell(row=row, column=1), font=Font(bold=True), fill=SUBHEADER_FILL, align=CENTER)
    _apply_style(ws.cell(row=row, column=2), fill=SUBHEADER_FILL, align=CENTER)
    
    for i, (date, _) in enumerate(days):
        scheds = Schedule.query.filter_by(store_id=store.id, date=date).all()
        emp_count = len(set(s.employee_id for s in scheds))
        cell = ws.cell(row=row, column=3 + i, value=emp_count)
        _apply_style(cell, font=Font(bold=True), fill=SUBHEADER_FILL, align=CENTER)
    _apply_style(ws.cell(row=row, column=3 + day_count), fill=SUBHEADER_FILL, align=CENTER)
    row += 1
    
    ws.cell(row=row, column=1, value='缺人时段')
    ws.cell(row=row, column=2, value='')
    _apply_style(ws.cell(row=row, column=1), font=Font(bold=True), align=CENTER)
    _apply_style(ws.cell(row=row, column=2), align=CENTER)
    
    for i, (date, _) in enumerate(days):
        scheds = Schedule.query.filter_by(store_id=store.id, date=date).all()
        slots = _time_slots(store, scheds)
        shortages = [s['time'] for s in slots if not s['meets_minimum']]
        cell_val = '、'.join(shortages) if shortages else '✓ 正常'
        cell = ws.cell(row=row, column=3 + i, value=cell_val)
        if shortages:
            _apply_style(cell, fill=WARN_FILL, align=CENTER)
        else:
            _apply_style(cell, fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'), align=CENTER)
    _apply_style(ws.cell(row=row, column=3 + day_count), align=CENTER)
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    for i in range(day_count):
        ws.column_dimensions[get_column_letter(3 + i)].width = 13
    ws.column_dimensions[get_column_letter(3 + day_count)].width = 10
    
    for r in range(4, row):
        ws.row_dimensions[r].height = 24

def get_store_summary_html(store, start_date_str, end_date_str):
    from scheduler import Scheduler
    start_date = _parse(start_date_str)
    end_date = _parse(end_date_str)
    scheduler = Scheduler(db)
    data = scheduler.get_store_view(store.id, start_date_str, end_date_str)
    if not data:
        return ''

    total_shortage_days = sum(1 for d in data['days'] if d['has_shortage'])
    total_staff = sum(d['total_staff'] for d in data['days'])
    total_senior = sum(d['senior_count'] for d in data['days'])
    total_hours = 0
    for d in data['days']:
        for s in d['schedules']:
            total_hours += s['duration']

    rows_html = ''
    for d in data['days']:
        shortage_html = ''
        if d['has_shortage']:
            shortage_html = f'<span style="color:#ff4d4f;font-weight:bold;">缺人时段: {", ".join(d["shortage_slots"])}</span>'
        else:
            shortage_html = '<span style="color:#52c41a;">✓ 排班完整</span>'

        staff_names = '、'.join([
            f"{s['employee_name']}{'[高]' if s['employee_skill'] == '高级' else ''}" 
            for s in d['schedules']
        ]) or '-'

        rows_html += f'''
        <tr>
            <td style="border:1px solid #ddd;padding:6px 8px;text-align:center;">{d['date']}</td>
            <td style="border:1px solid #ddd;padding:6px 8px;text-align:center;">{d['day_name']}</td>
            <td style="border:1px solid #ddd;padding:6px 8px;text-align:center;">{d['total_staff']}</td>
            <td style="border:1px solid #ddd;padding:6px 8px;text-align:center;">{d['senior_count']}</td>
            <td style="border:1px solid #ddd;padding:6px 8px;font-size:12px;">{staff_names}</td>
            <td style="border:1px solid #ddd;padding:6px 8px;text-align:center;">{shortage_html}</td>
        </tr>
        '''

    html = f'''
    <div style="font-family:Microsoft YaHei,Arial,sans-serif;max-width:800px;margin:0 auto;">
        <div style="background:linear-gradient(135deg,#667eea,#764ba2);color:white;padding:18px 24px;border-radius:6px;">
            <h2 style="margin:0;">{store.name} 排班通知</h2>
            <p style="margin:8px 0 0 0;opacity:0.9;">{start_date.isoformat()} 至 {end_date.isoformat()}</p>
        </div>

        <div style="background:#f5f7fa;padding:16px;margin:16px 0;border-radius:6px;">
            <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:12px;text-align:center;">
                <div>
                    <div style="font-size:22px;font-weight:bold;color:#667eea;">{(end_date - start_date).days + 1}</div>
                    <div style="font-size:12px;color:#888;">排班天数</div>
                </div>
                <div>
                    <div style="font-size:22px;font-weight:bold;color:#52c41a;">{(end_date - start_date).days + 1 - total_shortage_days}</div>
                    <div style="font-size:12px;color:#888;">完整天数</div>
                </div>
                <div>
                    <div style="font-size:22px;font-weight:bold;color:#fa8c16;">{round(total_hours, 1)}</div>
                    <div style="font-size:12px;color:#888;">总工时(h)</div>
                </div>
                <div>
                    <div style="font-size:22px;font-weight:bold;color:#ff4d4f;">{total_shortage_days}</div>
                    <div style="font-size:12px;color:#888;">缺人天数</div>
                </div>
            </div>
        </div>

        <table style="width:100%;border-collapse:collapse;margin:12px 0;">
            <thead>
                <tr style="background:#4472C4;color:white;">
                    <th style="border:1px solid #ddd;padding:8px;">日期</th>
                    <th style="border:1px solid #ddd;padding:8px;">星期</th>
                    <th style="border:1px solid #ddd;padding:8px;">人数</th>
                    <th style="border:1px solid #ddd;padding:8px;">高级</th>
                    <th style="border:1px solid #ddd;padding:8px;">在岗人员</th>
                    <th style="border:1px solid #ddd;padding:8px;">状态</th>
                </tr>
            </thead>
            <tbody>
                {rows_html}
            </tbody>
        </table>

        <div style="color:#888;font-size:12px;padding:12px;background:#fafafa;border-radius:6px;margin-top:16px;text-align:center;">
            此邮件由排班系统自动发送，如有疑问请联系人力部门。
        </div>
    </div>
    '''
    return html
